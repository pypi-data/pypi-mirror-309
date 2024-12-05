# coding: utf-8
import sys
from enum import IntEnum, unique
from pg_environment import config as penv
from pg_configuration.define import *
from pg_common import SingletonBase, log_info, log_fatal_exit, can_be_variable, check_list_same_type, \
    get_filename
import os
import glob
import shutil
import json
import openpyxl

__all__ = ("Configuration", "CodeType")


@unique
class CodeType(IntEnum):
    JAVA = 1
    PYTHON = 2
    GO = 3
    TS = 4
    JS = 5


@unique
class CfgType(IntEnum):
    b = 1
    s = 2
    c = 3


@unique
class ValueType(IntEnum):
    p_int = 1
    p_str = 2
    int = 3
    long = 4
    double = 5
    float = 6
    string = 7
    bool = 8
    byte = 9
    short = 10
    map = 11
    list = 12


class _Configuration(SingletonBase):
    def __init__(self):
        self._code_dir = {
            "JAVA": "java",
            "PYTHON": "python"
        }
        self._data_dir = "cfg_data"
        self._excel_dir = "excel"
        self._bin_dir = "bin_data"
        _cfg = penv.get_conf(KEY_CONFIGURATION)
        if _cfg:
            if KEY_CONFIGURATION_EXCEL_DIR in _cfg and _cfg[KEY_CONFIGURATION_EXCEL_DIR]:
                self._excel_dir = _cfg[KEY_CONFIGURATION_EXCEL_DIR]
            if KEY_CONFIGURATION_CODE_DIR in _cfg and _cfg[KEY_CONFIGURATION_CODE_DIR]:
                self._code_dir = _cfg[KEY_CONFIGURATION_CODE_DIR]
            if KEY_CONFIGURATION_CFG_DIR in _cfg and _cfg[KEY_CONFIGURATION_CFG_DIR]:
                self._data_dir = _cfg[KEY_CONFIGURATION_CFG_DIR]
            if KEY_CONFIGURATION_BIN_DIR in _cfg and _cfg[KEY_CONFIGURATION_BIN_DIR]:
                self._bin_dir = _cfg[KEY_CONFIGURATION_BIN_DIR]
        if not self._excel_dir.startswith("/"):
            self._excel_dir = "%s/%s" % (penv.get_pwd(), self._excel_dir)

        if not self._bin_dir.startswith("/"):
            self._bin_dir = "%s/%s" % (penv.get_pwd(), self._bin_dir)

        if not self._data_dir.startswith("/"):
            self._data_dir = "%s/%s" % (penv.get_pwd(), self._data_dir)

    @classmethod
    def check_row_one(cls, _row):
        """
        第一行
        所有单元格不能为空，且必须是CfgType枚举中定义的类型之一
        :param _row:
        :return:
        """
        for _cell in _row:
            if not _cell.value:
                log_fatal_exit("row one missing cfg type.")
            elif _cell.value not in CfgType.__members__.keys():
                log_fatal_exit("row one cfg type error: %s" % (_cell.value,))

    @classmethod
    def check_row_two(cls, _row):
        """
        第二行
        所有单元格不能为空，且必须是ValueType枚举中定义的类型之一
        必须有且只有一个p_int，p_str两者之一
        :param _row:
        :return:
        """
        _exist_primary_key = False
        for _cell in _row:
            if not _cell.value:
                log_fatal_exit("row two missing value type.")
            elif _cell.value not in ValueType.__members__.keys():
                log_fatal_exit("row two value type error: %s" % (_cell.value,))

            if _cell.value in ['p_int', 'p_str']:
                if _exist_primary_key:
                    log_fatal_exit("duplicate primary key")
                _exist_primary_key = True

        if not _exist_primary_key:
            log_fatal_exit("primary key must exist")

    @classmethod
    def check_row_three(cls, _row):
        """
        第三行
        所有单元格不能为空，且值必须可以能定义为变量的字符串
        且不能存在重复定义的字段名
        :param _row:
        :return:
        """
        _exists = set([])
        for _cell in _row:
            if not _cell.value:
                log_fatal_exit("row three missing name definition")
            elif not can_be_variable(_cell.value):
                log_fatal_exit("row three must can be variable: %s" % (_cell.value,))

            if _cell.value in _exists:
                log_fatal_exit("row three exist duplicate field definition: %s" % (_cell.value,))

            _exists.add(_cell.value)

    @classmethod
    def check_col(cls, _sheet, _col):
        """
        检测某一列
        从第五行开始至最大行，每个单元格内的元素必须符合第二行定义的数据类型的规范
        :param _sheet:
        :param _col:
        :return:
        """
        _vt = _sheet.cell(row=2, column=_col)
        _field = _sheet.cell(row=3, column=_col)
        log_info("checking column: %d, type: %s, field name: %s" % (_col, _vt.value, _field.value))
        _vte = ValueType[_vt.value]
        if (_vte == ValueType.p_int or _vte == ValueType.p_str) and not _field.value == 'id':
            log_fatal_exit(f"type: {_vt.value}, must named id.")
        _set = set([])
        for _row in range(5, _sheet.max_row + 1):
            _cell = _sheet.cell(row=_row, column=_col)
            _val = '' if _cell.value is None else str(_cell.value).strip()

            if _vte == ValueType.int or _vte == ValueType.long \
                    or _vte == ValueType.float or _vte == ValueType.double \
                    or _vte == ValueType.bool or _vte == ValueType.p_str \
                    or _vte == ValueType.p_int or _vte == ValueType.byte \
                    or _vte == ValueType.short:
                if not _val:
                    # 基础数据类型必须配置值
                    log_fatal_exit("basic data type can not be null")

                if _vte == ValueType.p_int or _vte == ValueType.p_str:
                    # 主键字段不能重复
                    if _val in _set:
                        log_fatal_exit("duplicate primary key: %s" % (_val,))
                    _set.add(_val)

                if _vte == ValueType.bool:
                    # bool 类型必须是true，false，0，1之一
                    if _val.lower() not in ('true', 'false', '0', '1'):
                        log_fatal_exit("bool value must be one of 'true' or 'false' or '0' or '1'")

                elif _vte != ValueType.p_str:
                    try:
                        float(_val)
                    except ValueError:
                        log_fatal_exit("data needs to be numeric")

            if _vte == ValueType.map and _val:
                try:
                    _v = json.loads(_val)

                    if type(_v) != dict:
                        log_fatal_exit("needs map data")
                except json.JSONDecodeError:
                    log_fatal_exit("not right map data")

            if _vte == ValueType.list and _val:
                try:
                    _v = json.loads(_val)

                    if type(_v) != list:
                        log_fatal_exit("needs list data")

                    if not check_list_same_type(_v):
                        log_fatal_exit("all values are not the same type")
                except json.JSONDecodeError:
                    log_fatal_exit("not right list data")

    @classmethod
    def check_single_sheet(cls, _sheet):
        log_info("checking sheet: %s" % (_sheet.title,))
        if _sheet.min_row > 1 or _sheet.min_column > 1:
            log_fatal_exit("every sheet needs start from cell A1")
        if _sheet.max_row <= 4:
            log_fatal_exit("every sheet needs 5 rows at least")
        _Configuration.check_row_one(_sheet[1])
        _Configuration.check_row_two(_sheet[2])
        _Configuration.check_row_three(_sheet[3])
        for _col in range(_sheet.min_column, _sheet.max_column + 1):
            _Configuration.check_col(_sheet, _col)

    def _export_json(self, _sheet, _fname):
        if not os.path.isdir(self._data_dir):
            os.mkdir(self._data_dir)
        _out = os.sep.join([self._data_dir, "%s%s.json" % (_fname, _sheet.title)])
        _list = []
        for _row in range(5, _sheet.max_row+1):
            _obj = {}
            _list.append(_obj)
            for _col in range(1, _sheet.max_column+1):
                _field = dict()
                _obj[_sheet.cell(row=3, column=_col).value] = _field
                _field['export'] = _sheet.cell(row=1, column=_col).value
                _field['type'] = _sheet.cell(row=2, column=_col).value
                _field['desc'] = _sheet.cell(row=4, column=_col).value
                # _field['value'] = _sheet.cell(row=_row, column=_col).value
                _val = _sheet.cell(row=_row, column=_col).value
                _vt = ValueType[_field['type']]
                if _vt == ValueType.map or _vt == ValueType.list:
                    if _val:
                        _field['value'] = json.loads(_val)
                    else:
                        if _vt == ValueType.map:
                            _field['value'] = {}
                        else:
                            _field['value'] = []
                elif _vt == ValueType.int or _vt == ValueType.byte or _vt == ValueType.p_int or _vt == ValueType.long or _vt == ValueType.short:
                    _field['value'] = int(_val)
                elif _vt == ValueType.float or _vt == ValueType.double:
                    _field['value'] = float(_val)
                elif _vt == ValueType.bool:
                    _field['value'] = True if _val in ['true', '1'] else False
                else:
                    if _val:
                        _field['value'] = _val
                    else:
                        _field['value'] = ''
        with open(_out, "w") as _fout:
            json.dump(_list, _fout, indent=2)
            log_info(f"export json file: {_out} success.")

    def check_single_file(self, _file, export_json=False):
        log_info("handle config file: %s" % (_file,))
        _wb = openpyxl.load_workbook(_file, read_only=True)
        _sheets = _wb.sheetnames
        if 'end' not in _sheets:
            log_fatal_exit("every excel config file needs an end sheet")
        elif 'end' == _sheets[0]:
            log_fatal_exit("there is no other sheet before end sheet")
        else:
            for _sheet in _sheets:
                if _sheet == "end":
                    break
                if not can_be_variable(_sheet):
                    log_fatal_exit("%s is not a suitable name for variable." % (_sheet,))
                _Configuration.check_single_sheet(_wb[_sheet])
                if export_json:
                    self._export_json(_wb[_sheet], get_filename(_file))

    def check_format(self, export_json=False):
        if os.path.isdir(self._excel_dir):
            _files = glob.glob(os.sep.join([self._excel_dir, "*.xlsx"]))
            # clean old files
            if os.path.isdir(self._data_dir):
                shutil.rmtree(self._data_dir)
            for _file in _files:
                _fname = get_filename(_file)
                if not can_be_variable(_fname):
                    log_fatal_exit("file name: %s must can be variable." % (_fname,))
                self.check_single_file(_file, export_json=export_json)
            log_info("all configuration files checked ok!")
        else:
            log_info("dir not exist: %s" % (self._excel_dir,))

    def export_data(self):
        self.check_format(export_json=True)

    def export_bin(self, make_compare=False):
        if make_compare:
            from pg_objectserialization import dumps
        else:
            from pg_common import dumps
        self.check_format(export_json=True)
        if os.path.isdir(self._bin_dir):
            shutil.rmtree(self._bin_dir)
        if not os.path.isdir(self._bin_dir):
            os.mkdir(self._bin_dir)

        _bin_client_dir = os.sep.join([self._bin_dir, "client"])
        os.mkdir(_bin_client_dir)
        _bin_server_dir = os.sep.join([self._bin_dir, "server"])
        os.mkdir(_bin_server_dir)

        _files = glob.glob(os.sep.join([self._data_dir, "*.json"]))
        for _file in _files:
            _fname = get_filename(_file)
            with open(_file, "r") as _fi:
                _cfg = json.load(_fi)
                _list_s = []
                _list_c = []
                for _item in _cfg:
                    _map_s = {}
                    _map_c = {}
                    for _k, _v in _item.items():
                        if _v['export'] == 's':
                            _map_s[_k] = _v['value']
                        elif _v['export'] == 'c':
                            _map_c[_k] = _v['value']
                        else:
                            _map_s[_k] = _v['value']
                            _map_c[_k] = _v['value']
                    if _map_c:
                        _list_c.append(_map_c)
                    if _map_s:
                        _list_s.append(_map_s)

                if _list_c:
                    _bin_c = dumps(_list_c)
                    _out_c = os.sep.join([self._bin_dir, "%s/%s.bin" % ("client", _fname, )])
                    with open(_out_c, "bw") as _foc:
                        _foc.write(_bin_c)
                        log_info(f"export bin file: {_out_c} success.")

                if _list_s:
                    _bin_s = dumps(_list_s)
                    _out_s = os.sep.join([self._bin_dir, "%s/%s.bin" % ("server", _fname,)])
                    with open(_out_s, "bw") as _fos:
                        _fos.write(_bin_s)
                        log_info(f"export bin file: {_out_s} success.")

    def export_python(self, _target_dir):
        _files = glob.glob(os.sep.join([self._data_dir, "*.json"]))
        _content=[]
        for _file in _files:
            _name = get_filename(_file)
            with open(_file, "r") as _fi:
                _cfg = json.load(_fi)
                if _cfg:
                    _c = "class %s(BaseModel):\n" % (_name, )
                    _data = _cfg[0]
                    _has_content = False
                    for _d in _data.keys():
                        if _data[_d]['export'] == 'c':
                            continue
                        _has_content = True
                        _t = _data[_d]['type']
                        if _t == 'p_int' or _t == 'long' or _t == 'byte' or _t == 'short':
                            _t = "int"
                        elif _t == 'p_str' or _t == 'string':
                            _t = "str"
                        elif _t == 'double':
                            _t = "float"
                        elif _t == 'map':
                            _t = "dict"
                        _c = _c + "\t%s: %s\n" % (_d, _t)
                    if _has_content:
                        _content.append(_c)

        _target_file = "%s/%s" % (_target_dir, "__init__.py")
        with open(_target_file, "w") as _fo:
            _out = "from pydantic import BaseModel\n"
            _fo.write("%s\n%s" % (_out, "\n\n".join(_content)))
            log_info(f"export python code: {_target_file} success.")


    def export_code(self, code_type:CodeType=CodeType.PYTHON):
        if code_type.name not in self._code_dir:
            log_fatal_exit(f"not config code dir for {code_type.name}.")
        _target_dir = self._code_dir[code_type.name]
        if os.path.isdir(_target_dir):
            shutil.rmtree(_target_dir)
        if not os.path.isdir(_target_dir):
            os.mkdir(_target_dir)

        if code_type==CodeType.PYTHON:
            self.export_python(_target_dir)


Configuration = _Configuration()
