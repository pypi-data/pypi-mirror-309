import pandas as pd
import base64
from typing import List
from openpyxl import load_workbook


def get_sheet_names(file_path: str, sheet_names_to_exclude: str = None):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    exclude_sheets = [sheet for sheet in sheet_names_to_exclude.split(",")]
    all_sheets = workbook.sheetnames
    included_sheets = [sheet for sheet in all_sheets if sheet not in exclude_sheets]
    return included_sheets


def get_excel_last_saved(path: str, dtype: str) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    last_saved = workbook.properties.modified
    if dtype == "string":
        last_saved = last_saved.strftime("%Y-%m-%dT%H:%M:%SZ")
    return last_saved


def get_excel_cell_value(path: str, sheet_name: str, cell: str) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook[sheet_name]
    return ws[cell].value


def get_excel_total_columns(path: str, sheet_name: str) -> int:
    # workbook = load_workbook(path, read_only=True, data_only=True)
    # ws = workbook[sheet_name]
    # return ws.max_column
    df = pd.read_excel(
        path,
        sheet_name,
    )
    return len(df.columns)


def get_excel_images(
    path: str,
    sheet_name: str,
) -> List[str]:
    workbook = load_workbook(path)
    ws = workbook[sheet_name]
    image_data_list = []

    for image in ws._images:
        img_data = image._data()
        base64_encoded_image = base64.b64encode(img_data).decode("utf-8")
        image_data_list.append(base64_encoded_image)

    return image_data_list


def get_column_letter(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def handle_usecols(path: str, sheet_name: str, usecols: str) -> str:
    """
    Handle usecols "A:B" or "A:~" to get all columns
    """
    splits = usecols.split(":")
    if len(splits) > 1 and splits[1] == "~":
        total_cols = get_excel_total_columns(path, sheet_name)
        max_column_letter = get_column_letter(total_cols)
        return f"{splits[0]}:{max_column_letter}"
    else:
        return usecols
