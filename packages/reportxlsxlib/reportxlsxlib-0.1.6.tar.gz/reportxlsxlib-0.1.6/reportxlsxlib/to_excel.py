# -*- coding:utf-8 -*- #

import ast
from datetime import datetime,timedelta,date
from openpyxl import Workbook as workbook,load_workbook
from copy import copy
import os
from pathlib import Path
import pkg_resources
from pythonnet import set_runtime
import platform
import inspect
import csv
import re

os_type = platform.system()
if os_type == "Windows":
    DLL_FILES = [pkg_resources.resource_filename('reportxlsxlib', f'resources/windows/{dll_name}') for dll_name in [
        'Spire.XLS.dll','SkiaSharp.dll'
    ]]
    Dll_CLASS = [pkg_resources.resource_filename('reportxlsxlib',f'resources/windows/{class_name}') for class_name in ['Spire.XLS','SkiaSharp']]
else:
    os.environ['PYTHONNET_PYDLL'] = '/usr/share/dotnet/shared/Microsoft.NETCore.App/6.0.30/System.Private.CoreLib.dll'
    os.environ['PYTHONNET_RUNTIME'] = 'coreclr'
    DLL_FILES = [pkg_resources.resource_filename('reportxlsxlib', f'resources/linux/{dll_name}') for dll_name in [
        'Spire.XLS.dll','SkiaSharp.dll'
    ]]
    Dll_CLASS = [pkg_resources.resource_filename('reportxlsxlib',f'resources/linux/{class_name}') for class_name in ['Spire.XLS','SkiaSharp']]

# if os_type == "Windows":
#     DLL_FILES = [pkg_resources.resource_filename('reportxlsxforlinux', f'resources/windows/{dll_name}') for dll_name in [
#         'Spire.XLS.dll','SkiaSharp.dll'
#     ]]
#     Dll_CLASS = [pkg_resources.resource_filename('reportxlsxforlinux',f'resources/windows/{class_name}') for class_name in ['Spire.XLS','SkiaSharp']]
# else:
#     os.environ['PYTHONNET_PYDLL'] = '/usr/share/dotnet/shared/Microsoft.NETCore.App/6.0.30/System.Private.CoreLib.dll'
#     os.environ['PYTHONNET_RUNTIME'] = 'coreclr'
#     DLL_FILES = [pkg_resources.resource_filename('reportxlsxforlinux', f'resources/linux/{dll_name}') for dll_name in [
#         'Spire.XLS.dll','SkiaSharp.dll'
#     ]]
#     Dll_CLASS = [pkg_resources.resource_filename('reportxlsxforlinux',f'resources/linux/{class_name}') for class_name in ['Spire.XLS','SkiaSharp']]


import clr
clr.AddReference(DLL_FILES[0])
#clr.AddReference(DLL_FILES[1])

from openpyxl.styles import numbers
from System.IO import *
#from SkiaSharp.SKColors import *
from Spire.Xls import *
from Spire.Xls.Core.Spreadsheet import HTMLOptions
from aspose.cells import HtmlSaveOptions, SaveFormat, Workbook as workbooks,Worksheet as worksheets,Cells as cellss
from aspose.cells.drawing import ImageType
from openpyxl.formatting.rule import ColorScaleRule
from System.Drawing import Color




class ColorHelper:
    # 定义更丰富的颜色映射
    COLOR_MAP = {
        "aliceblue": Color.AliceBlue, 
        'antiquewhite': Color.AntiqueWhite,
        'aqua': Color.Aqua,
        'aquamarine': Color.Aquamarine,
        'azure': Color.Azure,
        'beige': Color.Beige,
        'bisque': Color.Bisque,
        'black': Color.Black,
        'blanchedalmond': Color.BlanchedAlmond,
        'blue': Color.Blue,
        'blueviolet': Color.BlueViolet,
        'brown': Color.Brown,
        'burlywood': Color.BurlyWood,
        'cadetblue': Color.CadetBlue,
        'chartreuse': Color.Chartreuse,
        'chocolate': Color.Chocolate,
        'coral': Color.Coral,
        'cornflowerblue': Color.CornflowerBlue,
        'cornsilk': Color.Cornsilk,
        'crimson': Color.Crimson,
        'cyan': Color.Cyan,
        'darkblue': Color.DarkBlue,
        'darkcyan': Color.DarkCyan,
        'darkgoldenrod': Color.DarkGoldenrod,
        'darkgray': Color.DarkGray,
        'darkgreen': Color.DarkGreen,
        'darkkhaki': Color.DarkKhaki,
        'darkmagenta': Color.DarkMagenta,
        'darkolivegreen': Color.DarkOliveGreen,
        'darkorange': Color.DarkOrange,
        'darkorchid': Color.DarkOrchid,
        'darkred': Color.DarkRed,
        'darksalmon': Color.DarkSalmon,
        'darkseagreen': Color.DarkSeaGreen,
        'darkslateblue': Color.DarkSlateBlue,
        'darkslategray': Color.DarkSlateGray,
        'darkturquoise': Color.DarkTurquoise,
        'darkviolet': Color.DarkViolet,
        'deeppink': Color.DeepPink,
        'deepskyblue': Color.DeepSkyBlue,
        'dimgray': Color.DimGray,
        'dodgerblue': Color.DodgerBlue,
        'firebrick': Color.Firebrick,
        'floralwhite': Color.FloralWhite,
        'forestgreen': Color.ForestGreen,
        'fuchsia': Color.Fuchsia,
        'gainsboro': Color.Gainsboro,
        'ghostwhite': Color.GhostWhite,
        'gold': Color.Gold,
        'goldenrod': Color.Goldenrod,
        'gray': Color.Gray,
        'green': Color.Green,
        'greenyellow': Color.GreenYellow,
        'honeydew': Color.Honeydew,
        'hotpink': Color.HotPink,
        'indianred': Color.IndianRed,
        'indigo': Color.Indigo,
        'ivory': Color.Ivory,
        'khaki': Color.Khaki,
        'lavender': Color.Lavender,
        'lavenderblush': Color.LavenderBlush,
        'lawngreen': Color.LawnGreen,
        'lemonchiffon': Color.LemonChiffon,
        'lightblue': Color.LightBlue,
        'lightcoral': Color.LightCoral,
        'lightcyan': Color.LightCyan,
        'lightgoldenrodyellow': Color.LightGoldenrodYellow,
        'lightgreen': Color.LightGreen,
        'lightgray': Color.LightGray,
        'lightpink': Color.LightPink,
        'lightsalmon': Color.LightSalmon,
        'lightseagreen': Color.LightSeaGreen,
        'lightskyblue': Color.LightSkyBlue,
        'lightslategray': Color.LightSlateGray,
        'lightsteelblue': Color.LightSteelBlue,
        'lightyellow': Color.LightYellow,
        'lime': Color.Lime,
        'limegreen': Color.LimeGreen,
        'linen': Color.Linen,
        'magenta': Color.Magenta,
        'maroon': Color.Maroon,
        'mediumaquamarine': Color.MediumAquamarine,
        'mediumblue': Color.MediumBlue,
        'mediumorchid': Color.MediumOrchid,
        'mediumpurple': Color.MediumPurple,
        'mediumseagreen': Color.MediumSeaGreen,
        'mediumslateblue': Color.MediumSlateBlue,
        'mediumspringgreen': Color.MediumSpringGreen,
        'mediumturquoise': Color.MediumTurquoise,
        'mediumvioletred': Color.MediumVioletRed,
        'midnightblue': Color.MidnightBlue,
        'mintcream': Color.MintCream,
        'mistyrose': Color.MistyRose,
        'moccasin': Color.Moccasin,
        'navajowhite': Color.NavajoWhite,
        'navy': Color.Navy,
        'oldlace': Color.OldLace,
        'olive': Color.Olive,
        'olivedrab': Color.OliveDrab,
        'orange': Color.Orange,
        'orangered': Color.OrangeRed,
        'orchid': Color.Orchid,
        'palegoldenrod': Color.PaleGoldenrod,
        'palegreen': Color.PaleGreen,
        'paleturquoise': Color.PaleTurquoise,
        'palevioletred': Color.PaleVioletRed,
        'papayawhip': Color.PapayaWhip,
        'peachpuff': Color.PeachPuff,
        'peru': Color.Peru,
        'pink': Color.Pink,
        'plum': Color.Plum,
        'powderblue': Color.PowderBlue,
        'purple': Color.Purple,
        'red': Color.Red,
        'rosybrown': Color.RosyBrown,
        'royalblue': Color.RoyalBlue,
        'saddlebrown': Color.SaddleBrown,
        'salmon': Color.Salmon,
        'sandybrown': Color.SandyBrown,
        'seagreen': Color.SeaGreen,
        'seashell': Color.SeaShell,
        'sienna': Color.Sienna,
        'silver': Color.Silver,
        'skyblue': Color.SkyBlue,
        'slateblue': Color.SlateBlue,
        'slategray': Color.SlateGray,
        'snow': Color.Snow,
        'springgreen': Color.SpringGreen,
        'steelblue': Color.SteelBlue,
        'tan': Color.Tan,
        'teal': Color.Teal,
        'thistle': Color.Thistle,
        'tomato': Color.Tomato,
        'turquoise': Color.Turquoise,
        'violet': Color.Violet,
        'wheat': Color.Wheat,
        'white': Color.White,
        'whitesmoke': Color.WhiteSmoke,
        'yellow': Color.Yellow,
        'yellowgreen': Color.YellowGreen
    }

    @staticmethod
    def get_color_by_name(color_name):
        # 查找颜色并返回对应的颜色对象
                # 确保 color_name 是字符串
        if not isinstance(color_name, str):
            print("Invalid color name. Expected a string.")
            return None
        # 转换为小写并查找颜色
        color = ColorHelper.COLOR_MAP.get(color_name.lower())
        return color




class EmptyParameterError(Exception):
    """自定义异常类，用于处理参数为空的错误"""
    def __init__(self, parameter_name):
        self.parameter_name = parameter_name
        self.message = f"参数 '{parameter_name}' 不能为空"
        super().__init__(self.message)

class ParameterError(Exception):
    """自定义异常类，用于处理参数类型错误的情况"""
    def __init__(self, message, parameter_name):
        self.parameter_name = parameter_name
        self.message = message
        super().__init__(self.message)

def validate_parameters(params_to_check=None):
    if params_to_check is None:
        params_to_check = []

    def decorator(func):
        def wrapper(*args, **kwargs):
            signature = inspect.signature(func)
            bound_arguments = signature.bind(*args, **kwargs)
            bound_arguments.apply_defaults()
            
            for name, value in bound_arguments.arguments.items():
                if name == 'self':
                    continue
                
                # 只检查指定的参数是否为空
                if name in params_to_check:
                    if value is None or (isinstance(value, str) and not value.strip()):
                        raise EmptyParameterError(name)
                    
                # 参数类型检查
                param = signature.parameters[name]
                if param.annotation is not param.empty and not isinstance(value, param.annotation):
                    raise ParameterError(f"Parameter '{name}' is of type {type(value).__name__}, expected {param.annotation.__name__}", name)
            
            return func(*args, **kwargs)
        return wrapper
    return decorator





class sqlResult_to_Excel:

    @staticmethod
    @validate_parameters(params_to_check=['codilist','metalist','path','sheet_name'])

    def split_worksheet_by_column(file_path, sheet_index, target_column, output_directory, header_row_count):
        # Load the original workbook
        original_book = Workbook()
        original_book.LoadFromFile(file_path)

        # Get the specified worksheet
        sheet = original_book.Worksheets[sheet_index]

        # Get the header rows (based on specified header row count)
        header_row_end = header_row_count if sheet.LastRow >= header_row_count else sheet.LastRow
        header_rows = sheet.Range[1, 1, header_row_end, sheet.LastColumn]

        # Dictionary to store rows by their target column value
        rows_by_value = {}

        # Iterate through the rows to group them by the value in the target column
        for row in range(header_row_end + 1, sheet.LastRow + 1):
            cell_value = sheet.Range[row, target_column].Value
            cell_value_str = str(cell_value) if cell_value is not None else ""

            if cell_value_str not in rows_by_value:
                rows_by_value[cell_value_str] = []

            rows_by_value[cell_value_str].append([sheet.Range[row, col].Value if sheet.Range[row, col].Value is not None else "" for col in range(1, sheet.LastColumn + 1)])

        # Create a new workbook for each group and save it
        for value, rows in rows_by_value.items():
            # Create a copy of the original workbook
            new_book = Workbook()
            new_book.LoadFromFile(file_path)

            # Get the specified worksheet in the new workbook
            new_sheet = new_book.Worksheets[sheet_index]

            # Clear all rows except the header
            if new_sheet.LastRow > header_row_end:
                new_sheet.DeleteRow(header_row_end + 1, new_sheet.LastRow - header_row_end)

            # Copy the rows with the same value to the new workbook
            new_row_idx = header_row_end + 1
            for row in rows:
                for col_idx, cell_value in enumerate(row, start=1):
                    new_sheet.Range[new_row_idx, col_idx].Value = cell_value
                new_row_idx += 1

            # Set the column width in the new workbook
            for i in range(sheet.LastColumn):
                new_sheet.SetColumnWidth(i + 1, sheet.GetColumnWidth(i + 1))

            # Save the new workbook to a file named after the original file name + column value
            base_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_file_path = os.path.join(output_directory, f"{base_filename}_{value}.xlsx")
            new_book.SaveToFile(output_file_path, ExcelVersion.Version2016)
    

    @staticmethod
    @validate_parameters(params_to_check=['codilist','metalist','path','sheet_name'])
    def split_worksheet_by_column_with_style(file_path, sheet_index, target_column, output_directory, header_row_count):
        # Load the original workbook
        original_book = Workbook()
        original_book.LoadFromFile(file_path)

        # Get the specified worksheet
        sheet = original_book.Worksheets[sheet_index]

        # Get the header rows (based on specified header row count)
        header_row_end = header_row_count if sheet.LastRow >= header_row_count else sheet.LastRow
        header_rows = sheet.Range[1, 1, header_row_end, sheet.LastColumn]

        # Dictionary to store rows by their target column value
        rows_by_value = {}

        # Iterate through the rows to group them by the value in the target column
        for row in range(header_row_end + 1, sheet.LastRow + 1):
            cell_value = sheet.Range[row, target_column].Value
            cell_value_str = str(cell_value) if cell_value is not None else ""

            if cell_value_str not in rows_by_value:
                rows_by_value[cell_value_str] = []

            rows_by_value[cell_value_str].append([sheet.Range[row, col] for col in range(1, sheet.LastColumn + 1)])

        # Create a new workbook for each group and save it
        for value, rows in rows_by_value.items():
            # Create a copy of the original workbook
            new_book = Workbook()
            new_book.LoadFromFile(file_path)

            # Get the specified worksheet in the new workbook
            new_sheet = new_book.Worksheets[sheet_index]

            # Clear all rows except the header
            if new_sheet.LastRow > header_row_end:
                new_sheet.DeleteRow(header_row_end + 1, new_sheet.LastRow - header_row_end)

            # Copy the rows with the same value to the new workbook, maintaining formatting
            new_row_idx = header_row_end + 1
            for row in rows:
                for col_idx, cell in enumerate(row, start=1):
                    new_sheet.Range[new_row_idx, col_idx].Value = cell.Value
                    new_sheet.Range[new_row_idx, col_idx].Style = cell.Style
                new_row_idx += 1

            # Set the column width in the new workbook
            for i in range(sheet.LastColumn):
                new_sheet.SetColumnWidth(i + 1, sheet.GetColumnWidth(i + 1))

            # Save the new workbook to a file named after the original file name + column value
            base_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_file_path = os.path.join(output_directory, f"{base_filename}_{value}.xlsx")
            new_book.SaveToFile(output_file_path, ExcelVersion.Version2016)
    

    
    

    

    
    

    
    
    

    def __init__(self,
                 pic_dpi = 180,
                 font_path_for_linux=['/usr/share/fonts/chinese'],
                 *args
                 ):
        self.pic_dpi = pic_dpi
        self.font_path = font_path_for_linux
        self.rule1 = ColorScaleRule(
            start_type='min', start_color='FFF46A6A',  # 浅红色
            mid_type='percentile', mid_value=50, mid_color='FFFDEB83',  # 浅黄色
            end_type='max', end_color='FF63BE7B'  # 浅绿色
        )
        self.rule2 = ColorScaleRule(
            start_type='min', start_color='FF63BE7B',  # 浅绿色
            mid_type='percentile', mid_value=50, mid_color='FFFDEB83',  # 浅黄色
            end_type='max', end_color='FFF46A6A'  # 浅红色
        )
    
    def _copy_rows(self, source_ws, target_ws, num_rows):
        for row in source_ws.iter_rows(min_row=1, max_row=num_rows):
            new_row = [cell.value for cell in row]
            target_ws.append(new_row)
            for idx, cell in enumerate(row):
                new_cell = target_ws.cell(row=target_ws.max_row, column=idx + 1)
                self._copy_cell_format(cell, new_cell)
            target_ws.row_dimensions[target_ws.max_row].height = source_ws.row_dimensions[cell.row].height
    
    def _copy_merged_cells(self, source_ws, target_ws):
        # 复制合并单元格
        for merged_cell in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_cell))
            min_col, min_row, max_col, max_row = merged_cell.bounds
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    src_cell = source_ws.cell(row=row, column=col)
                    tgt_cell = target_ws.cell(row=row, column=col)
                    self._copy_cell_format(src_cell, tgt_cell)
            target_ws.row_dimensions[min_row].height = source_ws.row_dimensions[min_row].height
    
    def _copy_cell_format(self, source_cell, target_cell):
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)

    

    def _replace_text_in_html(self, file_path, new_text):
        try:
            # 读取HTML文件内容
            with open(file_path, 'r', encoding='utf-8') as file:
                html_content = file.read()
    
            # 正则表达式匹配 "Evaluation Only." 开头 "Aspose Pty Ltd" 结尾的文本
            pattern = r"Evaluation Only\..*?Aspose Pty Ltd"
    
            # 替换指定的文本
            updated_html_content = re.sub(pattern, new_text, html_content)
    
            print("文本替换成功！")
            return updated_html_content
        except Exception as e:
            print(f"出现错误: {e}")



    def _copy_sheet(self,source_ws, target_ws):
    # 复制数据和样式
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        # 处理合并单元格
        for merge_cell in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merge_cell))
    @validate_parameters(params_to_check=['codilist','metalist','path','sheet_name'])
    def update_by_template(self,codilist:list,metalist:list,path='',dis ='',sheet_name = '')->None:
        #try:
            #wb = load_workbook(path)
        #except Exception as e:
            #tmp_path = path.split('.')[0]+'_tmp.xlsx'
            #self._solve_out_of_index_error(path=path,dis=tmp_path)
            #wb = load_workbook(tmp_path)
            #wb = self._del_the_specific_sheet_stream(wb,del_str='Eva')
            #self._delete_file(tmp_path)
        #finally:
            #sh_name = sheet_name
            #if sh_name in wb.sheetnames:
                #ws = wb[sheet_name]
            #else:
                #ws = wb.create_sheet(title= sheet_name)
            #for item1,item2 in zip(codilist,metalist):
                #start_row = item1[0]
                #start_col = item1[1]
                #for r_idx,row in enumerate(item2):
                    ##print(row),print(type(row))
                    #for c_idx,value in enumerate(ast.literal_eval(row)):
                        ##print(value)
                        #ws.cell(row=start_row+r_idx,column=start_col+c_idx,value = value)
                        ##print(f'row:{start_row+r_idx}',f'column:{start_col+c_idx}')
                        ##print(ws.cell(row=start_row+r_idx,column=start_col+c_idx,value = value).value)
            #if dis:
                #wb.save(dis)
            #else:
                #wb.save(path)
        
        # 创建Workbook对象并加载Excel文件
        workbook = Workbook()
        workbook.LoadFromFile(path)

        # 获取指定名称的工作表
        sheet = workbook.Worksheets[sheet_name]

        # 确保位置列表和数据列表长度相同
        if len(codilist) != len(metalist):
            raise ValueError("Positions list and data blocks list must have the same length.")

        # 遍历位置和对应的数据块
        for pos, data_block in zip(codilist, metalist):
            # 获取起始单元格的行号和列号
            start_cell = sheet[pos]
            start_row, start_column = start_cell.Row, start_cell.Column
            # 填充数据块
            for i, row_data in enumerate(data_block):
                row_data = eval(row_data)
                if isinstance(row_data, tuple):
                    for j, value in enumerate(row_data):
                        # 计算目标单元格的位置
                        target_cell = sheet[start_row + i, start_column + j]
                        if isinstance(value,int):

                            target_cell.Value = str(value)
                        else:
                            target_cell.Value = str(value)
                else:
                    # 如果 row_data 不是元组，直接将其放入起始单元格
                    target_cell = sheet[start_row + i, start_column]
                    target_cell.Value = value

        # 保存更改后的文件

        if dis:
            workbook.SaveToFile(dis,FileFormat.Version2013)
        else:
            workbook.SaveToFile(path,FileFormat.Version2013)

        

        
    
    @validate_parameters(params_to_check=['codilist','metalist','path','target_path','sheet_name'])
    def ajust_and_copy_excel(self,source_codilist:list,target_codilist:list,path='',dis = '',sheet_name = ''):
        work_bk = workbooks(path)
        work_bk.calculate_formula()
        work_bk.save(path)
        wb = load_workbook(path,data_only=True)
        sheet_names = wb.sheetnames
        # 打印原始工作表名称
        print("原始工作表名称:", sheet_names)
        # 找到名称中包含 "Eva" 的工作表
        sheets_to_delete = [sheet for sheet in sheet_names if "Eva" in sheet]
        # 删除这些工作表
        if len(sheets_to_delete) > 0:
            for sheet in sheets_to_delete:
                std = wb[sheet]
                wb.remove(std)

        sheet = wb[sheet_name]
        #针对所有坐标开展排序工作
        for s_codi,t_codi in zip(source_codilist,target_codilist):
            source_range = sheet[s_codi]
            target = sheet[t_codi]
            start_row = target.row
            start_col = target.column
            for source_row_index,source_row in enumerate(source_range):
                for source_col_index,source_cell in enumerate(source_row):
                    target_row = start_row + source_row_index
                    target_col = start_col + source_col_index
                    #print(target_row,target_col)
                    target_cell = sheet.cell(row=target_row,column=target_col)
                    target_cell.value = source_cell.value
        if dis:
            wb.save(dis)         
        else:
            wb.save(path)

    @validate_parameters(params_to_check=['sort_range_list','sort_col_index_list','path','sheet_name'])
    def sort_excel(self,sort_range_list:list,sort_col_index_list:list,path='',dis = '',if_reverse = True,sheet_name = ''):
        try:
            wb = load_workbook(path,data_only=True)
        except Exception as e:
            tmp_path = path.split('.')[0]+'_tmp.xlsx'
            self._solve_out_of_index_error(path=path,dis=tmp_path)
            wb = load_workbook(tmp_path)
            wb = self._del_the_specific_sheet_stream(wb,del_str='Eva') 
            self._delete_file(tmp_path)
        finally:
            sheet = wb[sheet_name]
            true_sort_index_list = [x-1 for x in sort_col_index_list]
            for s_range ,sort_index in zip(sort_range_list,true_sort_index_list):
                sort_range = sheet[s_range]
                data = list(sort_range)
                data_all=[]
                for index,row in enumerate(data):
                    data_row =[]
                    for item in enumerate(row):
                        data_row.append(item[1].value)
                    data_all.append(data_row)
                data_all.sort(key = lambda row:row[sort_index],reverse=if_reverse)
        
                start_row = sort_range[0][0].row  # 范围的起始行号
                start_col = sort_range[0][0].column  # 范围的起始列号

                for row_index, row_data in enumerate(data_all):
                    for col_index,value in enumerate(row_data):
                        # 计算目标单元格的行和列位置
                        target_row = start_row + row_index
                        target_col = start_col + col_index
                        # 将排序后的值写入目标单元格
                        sheet.cell(row=target_row, column=target_col).value = value
            if dis:
                wb.save(dis)
            else:
                wb.save(path)

    @validate_parameters(params_to_check=['generate_range','path','sheet_name'])
    def excel_to_html(self,generate_range:tuple,path = '',sheet_name = '')->str:
        wb = Workbook()
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        old_sheet = source_wb.Worksheets[sheet_name] 
        # 选 
        cell_range = old_sheet.Range[generate_range] 
        sheet = wb.Worksheets[0]
        cell_range.Copy(sheet.Range[generate_range])
        
        stream = MemoryStream()
        options = HTMLOptions()
        options.SavedAsFragment = True
        options.isExportStyle = True
        sheet.SaveToHtml(stream,options)
        
        stream.Position = 0;
        reader = StreamReader(stream);
        htmlString = reader.ReadToEnd();
        wb.Dispose()
        return htmlString



    @validate_parameters(params_to_check=['starting_col_index','need_delete_cols','path','sheet_name'])
    def del_col_excl(self,starting_col_index:int,need_delete_cols:int,path='',dis= '',sheet_name ='')->None:
        #try:
            #wb = load_workbook(path)
        #except Exception as e:
            #tmp_path = path.split('.')[0]+'_tmp.xlsx'
            #self._solve_out_of_index_error(path=path,dis=tmp_path)
            #wb = load_workbook(tmp_path)
            #wb = self._del_the_specific_sheet_stream(wb,del_str='Eva') 
            #self._delete_file(tmp_path)
        #finally:
            #sh_name = sheet_name
            #if sh_name in wb.sheetnames:
                #ws = wb[sh_name]
            #starting_cols_index = starting_col_index
            #num_of_cols_to_delete = need_delete_cols
            #ws.delete_cols(starting_cols_index,num_of_cols_to_delete)
            #if dis:
                #wb.save(dis)
            #else:
                #wb.save(path)
        
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        del_index = []

        for i in range(source_wb.Worksheets.Count):
            print(source_wb.Worksheets[i].Name)
            if "Eva" in source_wb.Worksheets[i].Name:
                del_index.append(i)
        
        sheet = source_wb.Worksheets[sheet_name]

        for _ in range(need_delete_cols):
            sheet.DeleteColumn(starting_col_index)
        

        if dis:
            source_wb.SaveToFile(dis,FileFormat.Version2013)
        else:
            source_wb.SaveToFile(path,FileFormat.Version2013)

    @validate_parameters(params_to_check=['results_df_list','name_lists','dis'])
    def add_detail(self,results_df_list:list,name_lists:list,path='',dis = ''):
        if not os.path.exists(path) or not path:
            wb = workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
        else:
            try:
                wb = load_workbook(path)
            except Exception as e:
                tmp_path = path.split('.')[0]+'_tmp.xlsx'
                self._solve_out_of_index_error(path=path,dis=tmp_path)
                wb = load_workbook(tmp_path)
                wb = self._del_the_specific_sheet_stream(wb,del_str='Eva')
                self._delete_file(tmp_path)
                
        for item,name in zip(results_df_list,name_lists):
            new_sheet = wb.create_sheet(title=name)
            for row in item:
                str_row = tuple(str(item) for item in row)
                new_sheet.append(str_row)
        if not dis:
            wb.save(path)
        else:
            wb.save(dis)

    @validate_parameters(params_to_check=['sheet_need_change_list','need_set_col_tuple','path'])
    #设置查看参数
    def set_up_style_of_the_workbook(self,sheet_need_change_list:list,need_set_col_tuple:list,path='',dis = ''):
        try:
            wb = load_workbook(path)
        except Exception as e:

            tmp_path = path.split('.')[0]+'_tmp.xlsx'
            self._solve_out_of_index_error(path=path,dis=tmp_path)
            wb = load_workbook(tmp_path)
            wb = self._del_the_specific_sheet_stream(wb,del_str='Eva')
            self._delete_file(tmp_path)

        finally:
            for sheet,need_row in zip(sheet_need_change_list,need_set_col_tuple):
                ws = wb[sheet]  
                max_row = ws.max_row
                max_col = ws.max_column

                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        text_value = cell.value
                        try:
                            numeric_value = float(text_value)
                            cell.value = numeric_value
                        except (ValueError, TypeError):
                            pass
                print(need_row)
                for column in need_row:
                    print(need_row)
                    for row in range(1, max_row + 1):
                        cell = ws.cell(row=row, column=column)
                        cell.number_format = numbers.FORMAT_PERCENTAGE
            if dis:
                wb.save(dis)
            else:
                wb.save(path)

    @validate_parameters(params_to_check=['target_excel'])
    def split_the_excel(self, target_excel, sheet_options=None, format_setting=None, if_value_only=True):
        try:
            wb = load_workbook(target_excel)
        except Exception as e:
            tmp_path = target_excel.split('.')[0]+'_tmp.xlsx'
            self._solve_out_of_index_error(path=target_excel,dis=tmp_path)
            wb = load_workbook(tmp_path)
            wb = self._del_the_specific_sheet_stream(wb,del_str='Eva')
            self._delete_file(tmp_path)
        finally:
            wb_name = target_excel.split('.')[0]
            # 加载带数据的工作簿
            #wb_dataonly = load_workbook(target_excel, data_only=if_value_only)
        
            # 如果没有提供 sheet_options，默认处理所有工作表并进行拆分
            if sheet_options is None:
                sheet_options = {sheet: {'should_split': True, 'column_index': 1, 'header_row': 2} for sheet in wb.sheetnames}
        
            # 创建一个字典来存储所有工作表的拆分数据
            split_data = {}

            # 遍历所有工作表，根据需要进行处理
            for sheet_name, options in sheet_options.items():
                should_split = options.get('should_split', True)
                column_index = options.get('column_index', 1)
                header_row = options.get('header_row', 1)
                ws = wb[sheet_name]
            
                if should_split:
                    # 遍历每一行，根据指定字段的值进行拆分
                    for row in ws.iter_rows(min_row=header_row + 1):
                        key = row[column_index - 1].value  # 由于列索引从1开始，所以需要减1
                        if key is not None and key != 0:
                            if key not in split_data:
                                split_data[key] = {sheet: [] for sheet in wb.sheetnames}
                            split_data[key][sheet_name].append(row)

            print(f'the split data is {split_data.keys()}')

            # 创建新的工作簿并将拆分后的数据写入
            for key, sheets_data in split_data.items():
                # 确保只有在有拆分数据的情况下才创建工作簿
                if not any(sheets_data[sheet] for sheet in sheet_options if sheet_options[sheet]['should_split']):
                    continue  # 跳过没有实际拆分数据的项
            
                filename = f"{wb_name}-{key}.xlsx"
                wb_new = workbook()
            
                for sheet_name in wb.sheetnames:
                    if sheet_name == 'Sheet':
                        ws_new = wb_new.active
                        ws_new.title = sheet_name
                    else:
                        ws_new = wb_new.create_sheet(sheet_name)
                    ws = wb[sheet_name]
                
                    # 复制表头
                    header_row = sheet_options[sheet_name]['header_row']
                    self._copy_rows(ws, ws_new, header_row)
                
                    # 复制数据
                    if sheet_options[sheet_name]['should_split']:
                        for row_data in sheets_data[sheet_name]:
                            ws_new.append([cell.value for cell in row_data])
                            for idx, cell in enumerate(row_data):
                                new_cell = ws_new.cell(row=ws_new.max_row, column=idx + 1)
                                self._copy_cell_format(cell, new_cell)
                        
                    else:
                        for row in ws.iter_rows(min_row=header_row + 1):
                            ws_new.append([cell.value for cell in row])
                            if any(cell.value for cell in row):  # 检查是否有数据
                                for idx, cell in enumerate(row):
                                    new_cell = ws_new.cell(row=ws_new.max_row, column=idx + 1)
                                    self._copy_cell_format(cell, new_cell)
                        
                
                    # 复制合并单元格
                    self._copy_merged_cells(ws, ws_new)
            
                # 删除默认创建的工作表
                if "Sheet" in wb_new.sheetnames and "Sheet" not in wb.sheetnames:
                    del wb_new["Sheet"]
            
                # 设置单元条件格式 {'2级': {'A': 1, 'B': 2}}
                if format_setting:
                    for ws_item, solution in format_setting.items():
                        ws_c = wb_new[ws_item]
                        for col, rule_index in solution.items():
                            max_row = ws_c.max_row
                            rule = self.rule1 if rule_index == 1 else self.rule2
                            ws_c.conditional_formatting.add(f'{col}{header_row}:{col}{max_row}', rule)
                wb_new.save(filename)
        

    
    @validate_parameters(params_to_check=['target_wb','need_add_workbook','need_sheet_name'])
    def add_sheet(self,target_wb='',source_wb='',source_sheet='',new_sheet='new sheet'):
        #添加sheet
        try:
            wb = load_workbook(target_wb)
        except Exception as e:
            tmp_path = target_wb.split('.')[0]+'_tmp.xlsx'
            self._solve_out_of_index_error(path=target_wb,dis=tmp_path)
            wb = load_workbook(tmp_path)
            wb = self._del_the_specific_sheet_stream(wb,del_str='Eva')
            self._delete_file(tmp_path)

        try:
            wb_old = load_workbook(source_wb)
        except Exception as e:
            tmp_path = source_wb.split('.')[0]+'_tmp.xlsx'
            self._solve_out_of_index_error(path=source_wb,dis=tmp_path)
            wb_old = load_workbook(tmp_path)
            self._del_the_specific_sheet_stream(wb,del_str='Eva')
            self._delete_file(tmp_path)

        ws_old = wb_old[new_sheet]
        ws_new = wb.create_sheet(new_sheet)
        self._copy_sheet(ws_old,ws_new)
        wb.save(target_wb)
    
    @validate_parameters(params_to_check=['target_wb','source_workbook','source_sheet_name'])
    def add_sheet_with_style(self,target_wb='',source_workbook='',source_sheet_name='',new_sheet_name='new sheet'):
        #添加sheet
        wb = Workbook()
        wb.LoadFromFile(target_wb)
        if wb.Worksheets[new_sheet_name]:
            wb.Worksheets.Remove(wb.Worksheets[new_sheet_name])
        source_wb = Workbook()
        source_wb.LoadFromFile(source_workbook)
        old_sheet = source_wb.Worksheets[source_sheet_name] 
        wb.Worksheets.AddCopy(old_sheet)

        wb.Worksheets[source_sheet_name].Name = new_sheet_name
        new_sheet = wb.Worksheets[new_sheet_name]
        wb.ActiveSheetIndex = wb.Worksheets.Count - 1
        wb.SaveToFile(target_wb,FileFormat.Version2013)
    
    @validate_parameters(params_to_check=['source_file'])
    def convert_specific_sheet_to_html(self,source_file, target_file = '1.html'):
        # 加载 Excel 文件
        workbook = workbooks(source_file)
        # 指定转换的工作表
        #sheet_index = workbook.
        #workbook.getWorksheets().setActiveSheetIndex(sheet_index)
        # 创建 HtmlSaveOptions 对象
        options = HtmlSaveOptions()
        # 设置导出的 HTML 为单个页面
        options.export_active_worksheet_only = True
        # 将图像导出为 HTML 中的 base64 编码
        options.export_images_as_base64 = True
        options.exclude_unused_styles =True
        options.image_options.image_type = ImageType.PNG
        options.export_similar_border_style = True
        # 保持所有行和列的原始大小
        # 保存转换的 HTML 文件
        workbook.save(target_file, options)
        # 替换保护字体
        return self._replace_text_in_html(target_file,'')
        #Evaluation Only. Created with Aspose.Cells for Python via .NET. Copyright 2003 - 2024 Aspose Pty Ltd.

        
    @validate_parameters(params_to_check=['to_pic_generate_range','path'])
    def excel_to_pic(self,to_pic_generate_range,path = '',sheet_name = '通报',pic_name = 'results.png'):
        #生成新xlsx
        wb = Workbook()
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        old_sheet = source_wb.Worksheets[sheet_name] 
        wb.Worksheets.AddCopy(old_sheet)
        wb.Worksheets.RemoveAt(0)
        wb.Worksheets.RemoveAt(1)
       
        setting = ConverterSetting()
        setting.XDpi = self.pic_dpi
        setting.YDpi = self.pic_dpi
        wb.ConverterSetting =setting
        sheet = wb.Worksheets[sheet_name]
        
        os_type = platform.system()
        if os_type == "Windows":
            sheet.SaveToImage(pic_name,to_pic_generate_range[0],to_pic_generate_range[1],to_pic_generate_range[2],to_pic_generate_range[3] )
            wb.Dispose()
        elif os_type == "Linux":
            wb.CustomFontFileDirectory= self.font_path
            image_stream = sheet.ToImage(to_pic_generate_range[0],to_pic_generate_range[1],to_pic_generate_range[2],to_pic_generate_range[3])
            managedStream = SKManagedStream(image_stream)
            bitmap = SKBitmap.Decode(managedStream)
            image = SKImage.FromBitmap(bitmap)
            data = image.Encode(SKEncodedImageFormat.Png,self.pic_dpi)
            File.WriteAllBytes(pic_name,data.ToArray())
            wb.Dispose()
        else:
            print("不支持当前系统")

    @validate_parameters(params_to_check=['path','sheet_name'])
    def set_active_sheet(self,path = '', sheet_name ='', dis = ''):
    # 加载 Excel 文件
        # 创建Workbook对象并加载Excel文件
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        del_index = []

        for i in range(source_wb.Worksheets.Count):
            print(source_wb.Worksheets[i].Name)
            if "Eva" in source_wb.Worksheets[i].Name:
                del_index.append(i)
        
        sheet = source_wb.Worksheets[sheet_name]
        for index, sheet in enumerate(source_wb.Worksheets):
            if sheet.Name == sheet_name:
                source_wb.ActiveSheetIndex = index

        #source_wb.ActiveSheet = sheet

        if dis:
            source_wb.SaveToFile(dis,FileFormat.Version2013)
        else:
            source_wb.SaveToFile(path,FileFormat.Version2013)

    
    
    @validate_parameters(params_to_check=['path'])
    def caculate_excel_without_formula(self,path='',dis=''):
        work_bk = workbooks(path)
        work_bk.calculate_formula()
        #for i in range(work_bk.worksheets):
        for sheet in work_bk.worksheets:
            #sheet = worksheets[i]
            for cell in sheet.cells:
                if cell.is_formula:
                    calculated_value = cell.value
                    cell.put_value(calculated_value)
        work_bk.settings.formula_settings.enable_calculation_chain = False;
        work_bk.save(path)
        print('保存成功')
        
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        del_index = []
        for i in range(source_wb.Worksheets.Count):
            print(source_wb.Worksheets[i].Name)
            if "Eva" in source_wb.Worksheets[i].Name:
                del_index.append(i)
        print(del_index)

        for index in reversed(del_index):
            source_wb.Worksheets.RemoveAt(index)

        if dis:
            source_wb.SaveToFile(dis,FileFormat.Version2013)
        else:
            source_wb.SaveToFile(path,FileFormat.Version2013)
    
    @validate_parameters(params_to_check=['path'])
    def caculate_excel_with_formula(self,path='',dis=''):
        
        work_bk = workbooks(path)
        work_bk.calculate_formula()
        work_bk.save(path)

        print('保存成功')
        source_wb = Workbook()
        source_wb.LoadFromFile(path)
        del_index = []
        for i in range(source_wb.Worksheets.Count):
            print(source_wb.Worksheets[i].Name)
            if "Eva" in source_wb.Worksheets[i].Name:
                del_index.append(i)
        print(del_index)

        for index in reversed(del_index):
            source_wb.Worksheets.RemoveAt(index)

        if dis:
            source_wb.SaveToFile(dis,FileFormat.Version2013)
        else:
            source_wb.SaveToFile(path,FileFormat.Version2013)
        
    
    @validate_parameters(params_to_check=['path','sheet_name','range_address'])
    def delete_selected_range(self,path='',dis='',sheet_name='',range_address=''):
         # 创建 Workbook 对象并加载 Excel 文件
        workbook = Workbook()
        workbook.LoadFromFile(path)

        # 获取指定名称的工作表
        sheet = workbook.Worksheets[sheet_name]

        # 获取指定的单元格范围
        cell_range = sheet[range_address]

        # 清除范围内的内容
        cell_range.ClearContents()

        # 保存更改后的文件
        if dis:
            workbook.SaveToFile(dis,FileFormat.Version2013)
        else:
            workbook.SaveToFile(path,FileFormat.Version2013)
    


    @validate_parameters(params_to_check=['path','sheet_name','start_letter','columns_to_hide'])
    def hide_columns_by_letter(self,start_letter, columns_to_hide,path='',dis='',sheet_name=''):
        # 加载 Excel 文件
        workbook = Workbook()
        workbook.LoadFromFile(path)
    
        # 获取指定的工作表
        worksheet = workbook.Worksheets[sheet_name]
    
        # 将列字母转换为索引（1-based index）
        start_column_index = self._column_letter_to_index(start_letter)
    
        # 隐藏指定数量的列
        for i in range(start_column_index, start_column_index + columns_to_hide):
            worksheet.HideColumn(i)
    
        # 保存文件
        if dis:
            workbook.SaveToFile(dis,FileFormat.Version2013)
        else:
            workbook.SaveToFile(path,FileFormat.Version2013)
    
    def _column_letter_to_index(self,column_letter):
        """
        将Excel列字母（如 'A', 'B', 'AA', 'AAA'）转换为对应的列索引（1-based）。
        """
        column_letter = column_letter.upper()
        column_index = 0
        for char in column_letter:
            column_index = column_index * 26 + (ord(char) - ord('A') + 1)
        return column_index
    
    @validate_parameters(params_to_check=['path','source_sheet_name','source_range','target_sheet_name','target_start_cell'])
    def copy_range_within_workbook(self,source_sheet_name, source_range, target_sheet_name, target_start_cell,path,dis=''):
        # 加载 Excel 文件
        workbook = Workbook()
        workbook.LoadFromFile(path)
    
        # 获取源工作表和目标工作表
        source_sheet = workbook.Worksheets[source_sheet_name]
        target_sheet = workbook.Worksheets[target_sheet_name]
    
        # 获取源范围和目标起始位置的索引
        source_range = source_sheet[source_range]
        target_start_row, target_start_col = target_sheet[target_start_cell].Row, target_sheet[target_start_cell].Column
    
        # 遍历源范围，将内容复制到目标位置
        for i in range(source_range.Row, source_range.Row + source_range.RowCount):
            for j in range(source_range.Column, source_range.Column + source_range.ColumnCount):
                target_sheet[i - source_range.Row + target_start_row, j - source_range.Column + target_start_col].Value = source_sheet[i, j].Value
    
        # 保存文件

         # 保存文件
        if dis:
            workbook.SaveToFile(dis,FileFormat.Version2013)
        else:
            workbook.SaveToFile(path,FileFormat.Version2013)


   #def copy_range_within_workbook(self,source_sheet_name, source_range, target_sheet_name, target_start_cell,path,dis=''):
    def _parse_cell_position(self,cell):
        """
        将Excel单元格位置（如 'A1'）转换为行列数字索引。
    
        参数:
        - cell (str): 单元格地址，例如 'A1'
    
        返回:
        - tuple: (row, col) 行列元组（索引从 1 开始）
        """
        col = 0
        row = 0
        for i, char in enumerate(cell):
            if char.isalpha():
                col = col * 26 + (ord(char.upper()) - ord('A') + 1)
            else:
                row = int(cell[i:])  # 获取行数
                break
        return row, col

    @validate_parameters(params_to_check=['path','source_sheet_name','source_range','target_sheet_name','target_start_cell'])
        
    def update_cells(self,path,dis,sheet,start_cell: str, end_cell: str = None, value=None, formula=None):
        """
        更新指定单元格或范围的内容，支持更新文本、数值或公式。
    
        参数:
        - sheet (Worksheet): 要操作的工作表对象
        - start_cell (str): 开始单元格位置（例如 'A1'）
        - end_cell (str): 结束单元格位置（默认为 None，仅更新单个单元格）
        - value: 要插入的值（文本或数值，优先于公式）
        - formula: 要插入的公式（以 `=` 开头的字符串）
        """
        workbook = Workbook()
        workbook.LoadFromFile(path)
    
        # 获取源工作表和目标工作表
        source_sheet = workbook.Worksheets[sheet]

        # 解析单元格位置
        start_row, start_col = self._parse_cell_position(start_cell)
        if end_cell:
            end_row, end_col = self._parse_cell_position(end_cell)
        else:
            end_row, end_col = start_row, start_col  # 单个单元格
        # 遍历指定范围的单元格
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = source_sheet[row, col]
                # 更新内容
                if value is not None:
                    cell.Value = value
                elif formula:
                    cell.Formula = formula
        
                 # 保存文件
        if dis:
            workbook.SaveToFile(dis,FileFormat.Version2013)
        else:
            workbook.SaveToFile(path,FileFormat.Version2013)


    @validate_parameters(params_to_check=['path','sheet','col','target_values','target_color'])
    def set_row_color_by_column_values(self, path, dis, sheet, col, target_values, target_color):
        """
        通过匹配多个值设置整行颜色。
        """
        workbook = Workbook()
        workbook.LoadFromFile(path)
        trans_color = ColorHelper.get_color_by_name(target_color)
        # 获取源工作表
        source_sheet = workbook.Worksheets[sheet]
        target_values = [str(value) for value in target_values]

        for row in range(1, source_sheet.LastRow + 1):
            cell = source_sheet.Range[row, col]
            # 使用多个值匹配，如果单元格内容在目标值列表中，则设置颜色
            # print(cell.Value)
            if cell.Value is not None and str(cell.Value) in target_values:
                for col_index in range(1, source_sheet.LastColumn + 1):
                    source_sheet.Range[row, col_index].Style.Color = trans_color

        if dis:
            workbook.SaveToFile(dis, FileFormat.Version2013)
        else:
            workbook.SaveToFile(path, FileFormat.Version2013)

        
    @validate_parameters(params_to_check=['path','sheet','col','target_text','target_color'])
    def set_row_color_by_column_text(self,path,dis,sheet,col,target_text,target_color):
        """
        更新指定单元格或范围的内容，支持更新文本、数值或公式。
        """
        workbook = Workbook()
        workbook.LoadFromFile(path)
    
        # 获取源工作表和目标工作表
        source_sheet = workbook.Worksheets[sheet]
        trans_color = ColorHelper.get_color_by_name(target_color)

        for row in range(1,source_sheet.LastRow +1):
            cell = source_sheet.Range[row,col]
            if str(cell.Value) == str(target_text):
                for col_index in range(1,source_sheet.LastColumn +1):
                    source_sheet.Range[row,col_index].Style.Color = trans_color

        if dis:
            workbook.SaveToFile(dis,FileFormat.Version2013)
        else:
            workbook.SaveToFile(path,FileFormat.Version2013)

    @validate_parameters(params_to_check=['path','sheet','col','target_keyword','target_color'])
    def set_row_color_by_column_keyword(self, path, dis, sheet, col, target_keyword, target_color):
        """
        通过匹配关键字设置整行颜色。
        """
        workbook = Workbook()
        workbook.LoadFromFile(path)

        # 获取源工作表
        source_sheet = workbook.Worksheets[sheet]
        trans_color = ColorHelper.get_color_by_name(target_color)

        for row in range(1, source_sheet.LastRow + 1):
            cell = source_sheet.Range[row, col]
            # 使用关键字匹配，如果单元格内容包含目标关键字，则设置颜色
            if target_keyword in str(cell.Value):
                for col_index in range(1, source_sheet.LastColumn + 1):
                    source_sheet.Range[row, col_index].Style.Color = trans_color

        if dis:
            workbook.SaveToFile(dis, FileFormat.Version2013)
        else:
            workbook.SaveToFile(path, FileFormat.Version2013)

    # @validate_parameters()

    # def set_row_color_by_column_text(self,filePath, int sheetIndex, int column, string targetText, Color targetColor, string outputFilePath):
    #     Workbook workbook = new Workbook();
    #     workbook.LoadFromFile(filePath);

    #     Worksheet sheet = workbook.Worksheets[sheetIndex];

    #     for (int row = 1; row <= sheet.LastRow; row++)
    #     {
    #         CellRange cell = sheet.Range[row, column];
    #         if (cell.Text == targetText)
    #         {
    #             for (int col = 1; col <= sheet.LastColumn; col++)
    #             {
    #                 sheet.Range[row, col].Style.Color = targetColor;
    #             }
    #         }
    #     }

    #     workbook.SaveToFile(outputFilePath, ExcelVersion.Version2013);

    @validate_parameters(params_to_check=['path'])
    def _solve_out_of_index_error(self,path = '',dis = ''):
        wb = workbooks(path)
        wb.save(dis)
    
    @validate_parameters(params_to_check=['path','del_str'])
    def _del_the_specific_sheet(self,path='',del_str =''):
        # 找到名称中包含 "Eva" 的工作表
        wb = load_workbook(path)
        sheet_names = wb.sheetnames
        sheets_to_delete = [sheet for sheet in sheet_names if del_str in sheet]
        # 删除这些工作表
        if len(sheets_to_delete) > 0:
            for sheet in sheets_to_delete:
                std = wb[sheet]
                wb.remove(std)
        wb.save(path)
    def _del_the_specific_sheet_stream(self,wb:workbook,del_str ='')->workbook:
        # 找到名称中包含 "Eva" 的工作表
        sheet_names = wb.sheetnames
        sheets_to_delete = [sheet for sheet in sheet_names if del_str in sheet]
        # 删除这些工作表
        if len(sheets_to_delete) > 0:
            for sheet in sheets_to_delete:
                std = wb[sheet]
                wb.remove(std)
        return wb
    
    def _delete_file(self,file_path):
        try:
            file = Path(file_path)
            # 检查文件是否存在
            if file.exists() and file.is_file():
                file.unlink()
                print(f" '{file_path}' 文件删除成功.")
            else:
                print(f" '{file_path}' does not exist.")
        except Exception as e:
            print(f"文件删除出现异常: {e}")
    @validate_parameters(params_to_check=['path','sheet_name','target_column','keyword','range_start_col', 'range_end_col'])
    def delete_range_by_keyword(self,path, sheet_name, target_column, keyword, range_start_col, range_end_col, dis=None):
        """
        通过匹配关键字删除所在行中指定的单元格范围内容。
        """
        # Load the workbook
        workbook = Workbook()
        workbook.LoadFromFile(path)

        # Get the specified worksheet
        sheet = workbook.Worksheets[sheet_name]

        # Iterate through the rows to find the keyword and clear the specified range
        for row in range(1, sheet.LastRow + 1):
            cell_value = sheet.Range[row, target_column].Value
            if keyword in str(cell_value):
                cell_range = sheet.Range[row, range_start_col, row, range_end_col]
                cell_range.ClearContents()

        # Save the workbook
        if dis:
            workbook.SaveToFile(dis, FileFormat.Version2013)
        else:
            workbook.SaveToFile(path, FileFormat.Version2013)    



if __name__ =='__main__':
    excl = sqlResult_to_Excel()


    #excl.update_cells('template_v1.xlsx','template_v2.xlsx','二级','A100','B100',value='这是一个测试活动')
    ## excl.caculate_excel_with_formula('template_v2.xlsx')
    #excl.set_row_color_by_column_text('每日套餐包和促销受理统计_20241029.xlsx','template_test.xlsx','促销和可选包',6,'流量包','red')
    #excl.delete_range_by_keyword('每日套餐包和促销受理统计_20241029.xlsx','促销和可选包',4,'可选包',25,38,dis='templatetest.xlsx')
    #excl.set_row_color_by_column_values('每日套餐包和促销受理统计_20241029.xlsx','template_test.xlsx','促销和可选包',1,['552496185','838237','552474410'],Color.Yellow)
    #excl.set_row_color_by_column_keyword('每日套餐包和促销受理统计_20241029.xlsx','template_test.xlsx','促销和可选包',2,'流量',Color.Yellow)
    # sqlResult_to_Excel.split_worksheet_by_column_with_style('template_v2.xlsx','三级',1,'./',5)
    #excl.add_sheet_with_style('final_excel_v1.xlsx','template_new_v1.xlsx','通报','0833')
    #excl.caculate_excel('template_v1.xlsx','test.xlsx')

    #excl.del_col_excl(84,50,'test.xlsx',sheet_name='套餐')

    #excl.set_active_sheet('template_v1.xlsx',sheet_name='套餐')


    #def read_csv_file(file_path):
        #with open(file_path, 'r', newline='',encoding= 'utf-8') as csvfile:
            #data = [row for row in csv.reader(csvfile)]
        #return data
    #result_data = read_csv_file('results.csv')


    #codi_1 =['AR4']
    #codi_2 = ['AN4']

    #excl.update_by_template(
        #codi_1,
        #[result_data[0]],
        #path = 'template_v1.xlsx',
        ##dis= 'template_v2.xlsx',
        #sheet_name='套餐'
    #)

    #excl.update_by_template(
        #codi_2,
        #[result_data[1]],
        #path = 'template_v1.xlsx',
        #dis= 'template_v2.xlsx',
        #sheet_name='促销和可选包'
    #)
    #excl.caculate_excel_without_formula('template_v2.xlsx','template_v3.xlsx')
    #excl.del_col_excl(41,50,'template_v3.xlsx',sheet_name='套餐')
    #excl.del_col_excl(39,50,'template_v3.xlsx',sheet_name='促销和可选包')
    #excl.set_active_sheet('template_v3.xlsx',sheet_name='套餐')

    #excl.delete_selected_range('template_v3.xlsx',sheet_name='套餐',range_address='AH4:AO403')
    #excl.delete_selected_range('template_v3.xlsx',sheet_name='促销和可选包',range_address='AF4:AM1003')
    #excl.caculate_excel_v1('template_v2.xlsx','template_v3.xlsx')
    #excl.hide_columns_by_letter(start_letter='AO',columns_to_hide=5,path='template_v3.xlsx',dis='template_v4.xlsx',sheet_name='套餐')
    #excl.copy_range_within_workbook('套餐', 'A5:D18', '套餐', 'A500','template_v4.xlsx',dis='')
    #with open('3.html','w',encoding= 'utf-8') as file:
        #file.write(excl.convert_specific_sheet_to_html('价值管控通报--到门店.xlsx'))
