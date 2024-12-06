import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


def delete_file(output_path):
    if output_path.exists():
        logging.warning(f"Removing file{output_path} ")
        output_path.unlink()
    else:
        logging.warning(f"file {output_path} doesn't exist")


def _remove_nulls(df, columns):
    for column in columns:
        df[column] = df[column].fillna("")
        df[column] = df[column].replace("None", "")
    return df


def save_excel_file(df: pd.DataFrame, output_path: Path):
    # Check if DataFrame is empty
    if df.empty:
        logging.warning("Empty DataFrame, nothing to write")
        return

    # Set options to display the full DataFrame
    pd.set_option("display.max_rows", None)  # Show all rows
    pd.set_option("display.max_columns", None)  # Show all columns
    pd.set_option(
        "display.max_colwidth", None
    )  # Show full column content without truncation
    pd.set_option("display.width", 1000)  # Set display width to avoid line wrapping
    logging.warning(df.head())

    # Clean nulls in specific columns
    # df = _remove_nulls(df, ['upstream_transformation', 'downstream_transformation'])
    df = _remove_nulls(df, ["upstream_transformation"])

    # Check if the file already exists
    if output_path.exists():
        logging.warning(f"Appending to existing file: {output_path}")
        wb = load_workbook(output_path)
        ws = wb.active
        logging.debug(f"Number of rows already exists: {ws.max_row}")
        start_row = ws.max_row + 1  # Start appending from the next row
    else:
        # Create a new workbook if the file doesn't exist
        logging.warning(f"Creating new file: {output_path}")
        df.to_excel(output_path, index=False)  # Save initial data for headers
        wb = load_workbook(output_path)
        ws = wb.active
        start_row = 2  # Starting row for new data (after headers)

    # Append DataFrame rows to the worksheet
    for row_data in df.itertuples(index=False, name=None):
        ws.append(row_data)

    # Apply styles to headers
    header_fill = PatternFill(
        start_color="FFDD99", end_color="FFDD99", fill_type="solid"
    )  # Light orange for header
    for cell in ws[1]:  # First row as header
        cell.fill = header_fill
        cell.font = Font(bold=True)

    # Set specific column widths

    DATASET_COL = "A"
    COLNAME_COL = "B"
    UPSTREAM_COL = "C"
    # DOWNSTREAM_COL = "D"
    ws.column_dimensions[DATASET_COL].width = 30  # Column name
    ws.column_dimensions[COLNAME_COL].width = 30  # Column name
    ws.column_dimensions[UPSTREAM_COL].width = 60  # Upstream transformation
    # ws.column_dimensions[DOWNSTREAM_COL].width = 60  # Downstream transformation

    # Process rows to merge "Name", "Type", and "Nullable" columns and concatenate "Upstream" and "Downstream"
    current_row = start_row
    while current_row <= ws.max_row:
        dataset_name_value = ws[f"{DATASET_COL}{current_row}"].value
        column_name_value = ws[f"{COLNAME_COL}{current_row}"].value
        next_row = current_row + 1

        # Initialize concatenated values for "Upstream" and "Downstream"
        upstream_concat = ws[f"{UPSTREAM_COL}{current_row}"].value or ""
        # downstream_concat = ws[f"{DOWNSTREAM_COL}{current_row}"].value or ""

        # Check if the next rows have the same "Name" value
        while (
            next_row <= ws.max_row
            and ws[f"{DATASET_COL}{next_row}"].value == dataset_name_value
            and ws[f"{COLNAME_COL}{next_row}"].value == column_name_value
        ):
            # Concatenate "Upstream" and "Downstream" values
            if ws[f"{UPSTREAM_COL}{next_row}"].value:
                upstream_concat += f"\n{ws[f'{UPSTREAM_COL}{next_row}'].value}"
            # if ws[f"{DOWNSTREAM_COL}{next_row}"].value:
            #     downstream_concat += f"\n{ws[f'{DOWNSTREAM_COL}{next_row}'].value}"
            next_row += 1

        # Update "Upstream" and "Downstream" columns with concatenated values
        ws[f"{UPSTREAM_COL}{current_row}"].value = upstream_concat
        # ws[f"{DOWNSTREAM_COL}{current_row}"].value = downstream_concat
        ws[f"{UPSTREAM_COL}{current_row}"].alignment = Alignment(
            wrap_text=True, vertical="center"
        )
        # ws[f"{DOWNSTREAM_COL}{current_row}"].alignment = Alignment(wrap_text=True, vertical="center")

        # Merge cells if there are multiple rows with the same "Name"
        if next_row - current_row > 1:
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=next_row - 1,
                end_column=1,
            )  # Merge "Name" column
            ws.merge_cells(
                start_row=current_row,
                start_column=2,
                end_row=next_row - 1,
                end_column=2,
            )  # Merge "Type" column
            ws.merge_cells(
                start_row=current_row,
                start_column=3,
                end_row=next_row - 1,
                end_column=3,
            )  # Merge "Nullable" column

        # Move to the next distinct "Name" value
        current_row = next_row

    # Save changes
    wb.save(output_path)
    logging.warning(
        f"Excel file with merged cells and concatenated transformations saved to {output_path}"
    )
