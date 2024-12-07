from openpyxl import Workbook, load_workbook
from typing import Optional
import os


def get_parent_directory() -> str:
    """returns a string containing the absolute filepath to the
    parent directory

    Parameters
    ----------

    Returns
    -------
    string containing the absolute filepath
    """

    return os.path.abspath(os.path.join(os.getcwd(), os.pardir))


def load_file(filename: str) -> Workbook:
    """loads a file and returns a Workbook object

    Parameters
    ----------
    filename : str
        filename of the workbook to load

    Returns
    -------
    Workbook
    """

    try:
        wb = load_workbook(filename=filename)
    except Exception as err:
        print("could not load workbook ({}):\n{}".format(filename, err))

    return wb


def save_file(workbook: Workbook,
              filename: Optional[str] = "save.xlsx") -> None:
    """saves a workbook

    Parameters
    ----------
    workbook : Workbook
        openpyxl.Workbook object to save
    filename : Optional[str], default: "save.xlsx"
        String to use as the filename

    Returns
    -------
    None
    """

    try:
        workbook.save(filename="hello_world.xlsx")
    except TypeError as err:
        print("Type Error: could not save workbook:\n{}".format(err))
    except Exception as err:
        print("could not save workbook: {}".format(err))


def main():
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "hello"
    sheet["B1"] = "world!"

    save_file(workbook, "hello_world.xlsx")
    print("done!")


if __name__ == '__main__':
    main()
