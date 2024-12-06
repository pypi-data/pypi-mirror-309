import io
import json
import logging
from typing import List

import openpyxl
from fastapi import UploadFile
from sqlalchemy.orm import sessionmaker

from src.hubrdvmairie.models.meeting_point import MeetingPoint

from ..crud.crud_meeting_point import meetingPoint as crud

logger = logging.getLogger(__name__)


def get_all(session) -> List[MeetingPoint]:
    try:
        return crud.get_all(session)
    except Exception as e:
        logger.error("Error while getting all meeting points : %s", str(e))


def get_by_ugf(session, ugf: str) -> MeetingPoint:
    try:
        return crud.get_by_ugf(session, ugf=ugf)
    except Exception as e:
        logger.error("Error while getting meeting point by ugf : %s", str(e))


async def update_meeting_points_table(session: sessionmaker, uploaded_file: UploadFile):
    meeting_points = await read_meeting_point_from_file_streaming(uploaded_file)
    create_list = []
    unchanged_list = []
    updated_list = []
    nb_meeting_points = len(meeting_points)
    ligne = 1
    name_meeting = ""
    for meeting_point in meeting_points:
        name_meeting = meeting_point.editor_name_and_id
        try:
            print(
                meeting_point.ugf,
                meeting_point.editor_name_and_id,
                meeting_point.city_name,
                meeting_point.id_editor,
            )
            res = crud.save_or_update(session, obj_in=meeting_point)
            if res[0] == "created":
                create_list.append(res[1])
            elif res[0] == "updated":
                updated_list.append(res[1])
            else:
                unchanged_list.append(res[1])
            ligne += 1
        except Exception:
            print("ligne : " + str(ligne), "name_meeting : " + name_meeting)

    return json.dumps(
        {
            "nb_meeting_points": nb_meeting_points,
            "created : ": str(len(create_list)),
            "updated : ": str(len(updated_list)),
            "unchanged : ": str(len(unchanged_list)),
        }
    )


async def read_meeting_point_from_file_streaming(
    uploaded_file: UploadFile,
) -> MeetingPoint:
    # read file depending on its type
    if uploaded_file.filename.endswith(".xlsx"):
        return await read_meeting_points_file_streaming(uploaded_file)
    else:
        raise TypeError("Unknown file type : " + str(uploaded_file.filename))


async def read_meeting_points_file_streaming(uploaded_file: UploadFile):
    try:
        meeting_points = set()

        file_content = await uploaded_file.read()
        xls_data = io.BytesIO(file_content)
        wb = openpyxl.load_workbook(filename=xls_data)
        worksheet = wb.active

        for i in range(2, worksheet.max_row):
            ugf = str(worksheet.cell(i, 1).value)
            editor_name_and_id = worksheet.cell(i, 2).value
            city_name = worksheet.cell(i, 3).value
            id_editor = worksheet.cell(i, 4).value
            if ugf != "" or editor_name_and_id is not None or id_editor is not None:
                meeting_point = MeetingPoint(
                    ugf=ugf,
                    editor_name_and_id=editor_name_and_id,
                    city_name=city_name,
                    id_editor=id_editor,
                )
                meeting_points.add(meeting_point)
        wb.close()
        return meeting_points
    except Exception as e:
        print("reading : " + e)
