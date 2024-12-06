import io
import json
import logging
from collections import defaultdict
from typing import List, Set, Tuple

from fastapi import UploadFile
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy.orm import sessionmaker

from src.hubrdvmairie.core.custom_validation_error import CustomValidationError
from src.hubrdvmairie.schemas.certified_municipality import CertifiedMunicipality

from ..crud.crud_certified_municipality import certified_municipality as crud

logger = logging.getLogger(__name__)


def get_all(session) -> List[CertifiedMunicipality]:
    try:
        return crud.get_all(session)
    except Exception as e:
        logger.error("Error while getting all certified_municipality : %s", str(e))


def get_all_certified_municipalities(
    session: sessionmaker,
) -> List[CertifiedMunicipality]:
    municipalities = get_all(session)
    certified_municipalities_complete = []
    for certified_municipality in municipalities:
        complete = certified_municipality.complete()
        certified_municipalities_complete.append(complete)
    departments = defaultdict(list)

    for municipality in certified_municipalities_complete:
        department_code = municipality["department_code"]
        departments[department_code].append(municipality)

    departments_list = []
    for department_code, municipalities in departments.items():
        department_info = {
            "department_code": department_code,
            "department_name": municipalities[0]["department_name"],
            "municipalities": municipalities,
        }
        departments_list.append(department_info)
    departments_list = sorted(departments_list, key=lambda x: x["department_code"])
    return departments_list


async def update_certified_municipality_table(
    session: sessionmaker, uploaded_file: UploadFile
):
    (
        certified_municipalities,
        errors,
    ) = await read_certified_municipality_from_file_streaming(uploaded_file)
    if errors:
        raise CustomValidationError(errors)
    else:
        create_list = []
        unchanged_list = []
        nb_certified_municipalities = len(certified_municipalities)
        for certified_municipality in certified_municipalities:
            res = crud.save_or_update(session, obj_in=certified_municipality)
            if res[0] == "created":
                create_list.append(res[1])
            else:
                unchanged_list.append(res[1])

        response_data = {
            "nb_certified_municipality": nb_certified_municipalities,
            "created": str(len(create_list)),
            "unchanged": str(len(unchanged_list)),
        }
        return json.dumps(response_data)


async def read_certified_municipality_from_file_streaming(
    uploaded_file: UploadFile,
) -> CertifiedMunicipality:
    # read file depending on its type
    if uploaded_file.filename.endswith(".xlsx"):
        return await read_certified_municipality_file_streaming(uploaded_file)
    else:
        raise TypeError("Unknown file type : " + str(uploaded_file.filename))


async def read_certified_municipality_file_streaming(
    uploaded_file: UploadFile,
) -> Tuple[Set["CertifiedMunicipality"], list]:
    certified_municipalities = set()
    errors = []

    file_content = await uploaded_file.read()
    xls_data = io.BytesIO(file_content)
    data = load_workbook(xls_data, read_only=True).active

    for i, row in enumerate(
        data.iter_rows(min_row=2, max_row=data.max_row, values_only=True), start=2
    ):
        row_data = {
            "town_hall_name": row[0],
            "ugf": str(row[1]),
            "address": row[2],
            "zip_code": str(row[3]),
            "city_name": row[4],
            "phone_number": row[5],
            "website": row[6],
            "appointment_details": str(row[7]),
            "service_opening_date": str(row[8]),
            "label": row[9],
        }

        if all(value in ["None", None] for value in row_data.values()):
            errors.append(
                f"ligne n°{i} : {row} - erreur : Cette ligne est vide, merci de la supprimer"
            )
            continue

        if any(value in ["None", None] for value in row_data.values()):
            errors.append(f"ligne n°{i} : {row} - erreur : au moins un champ est vide")
            continue

        try:
            certified_municipality = CertifiedMunicipality(
                ugf=row_data["ugf"],
                town_hall_name=row_data["town_hall_name"],
                address=row_data["address"],
                zip_code=row_data["zip_code"],
                city_name=row_data["city_name"],
                phone_number=row_data["phone_number"],
                website=row_data["website"],
                appointment_details=row_data["appointment_details"],
                service_opening_date=row_data["service_opening_date"],
                label=row_data["label"],
            )
            CertifiedMunicipality.parse_obj(certified_municipality)
            certified_municipalities.add(certified_municipality)
        except ValidationError as e:
            logger.error(
                "Error while reading certified_municipality from file : %s", str(e)
            )
            errors.append(f"ligne n°{i} : {row} - erreur : {str(e)}")

    return certified_municipalities, errors
