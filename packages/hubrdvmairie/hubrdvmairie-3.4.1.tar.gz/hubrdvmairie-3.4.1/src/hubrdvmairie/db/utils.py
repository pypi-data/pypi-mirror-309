import os
from pathlib import Path
from typing import List

import openpyxl
from geopy import distance

from ..core.config import get_settings
from ..models.editor import Editor
from ..models.municipality import Municipality, OfflineMunicipality


def get_all_editors() -> List[Editor]:
    """
    Get global editors from the settings object
    """
    editors_stored: List[Editor] = get_settings().editors_list.copy()
    return editors_stored


def set_all_editors(editors: List[Editor]):
    """
    Set global editors list in the settings object
    """
    get_settings().editors_list = editors


def get_all_meeting_points() -> List[Municipality]:
    """
    Get global meeting point from the settings object
    """
    meeting_point_stored: List[Municipality] = get_settings().meeting_point_list.copy()
    return meeting_point_stored


def set_all_meeting_points(meeting_points: List[Municipality]):
    """
    Set global meeting points list in the settings object
    """
    get_settings().meeting_point_list = meeting_points

    if os.environ.get("EXCEL_PASSWORD"):
        offline_meeting_points = read_offline_meeting_points_file(meeting_points)
        set_all_offline_meeting_points(offline_meeting_points)


def get_ws_use_rates(ip_address) -> int:
    """
    Get global WebSocket use rates from the settings object for a specific ip address
    """
    if ip_address in get_settings().ws_use_rates:
        return get_settings().ws_use_rates[ip_address]
    else:
        return 0


def add_ws_use_rates(ip_address):
    """
    Add 1 to global websocket use rates list in the settings object for a specific ip address
    """
    if ip_address not in get_settings().ws_use_rates:
        get_settings().ws_use_rates[ip_address] = 0
    get_settings().ws_use_rates[ip_address] += 1


def reset_all_ws_use_rates():
    """
    Reset global websocket use rates list in the settings object
    """
    get_settings().ws_use_rates = {}


def read_offline_meeting_points_file(
    online_meeting_points,
) -> List[OfflineMunicipality]:
    """
    Read meeting points file and decrypt it
    """
    decrypted_offline_meeting_points = []
    folder_path = Path(__file__).parent
    file_rel_path = folder_path / Path("offline_meeting_points.xlsx")
    workbook = openpyxl.load_workbook(file_rel_path)
    main_sheet = workbook.worksheets[0]
    max_row = main_sheet.max_row

    meeting_point_id = 5000
    for i in range(2, max_row + 1):
        # if main_sheet.cell(row=i, column=1).value:
        if (
            main_sheet.cell(row=i, column=7).value
            and main_sheet.cell(row=i, column=8).value
        ):
            try:
                decrypted_offline_meeting_points.append(
                    {
                        "id": str(meeting_point_id),
                        "ugf": main_sheet.cell(row=i, column=1).value,
                        "municipality": main_sheet.cell(row=i, column=2).value,
                        "longitude": float(main_sheet.cell(row=i, column=7).value),
                        "latitude": float(main_sheet.cell(row=i, column=8).value),
                        "public_entry_address": (
                            main_sheet.cell(row=i, column=3).value or ""
                        )
                        + (main_sheet.cell(row=i, column=4).value or ""),
                        "zip_code": main_sheet.cell(row=i, column=5).value,
                        "decoded_city_name": main_sheet.cell(row=i, column=6).value,
                        "website": main_sheet.cell(row=i, column=11).value,
                        "logo": None,
                        "phone_number": main_sheet.cell(row=i, column=9).value,
                    }
                )
            except Exception as e:
                print("Erreur lors de la lecture des offline meeting points", str(e))
        meeting_point_id += 1
    # else:
    #    break

    filtered_offline_meeting_points = []
    for offline_meeting_point in decrypted_offline_meeting_points:
        is_online = False
        for online_meeting_point in online_meeting_points:
            if (
                str(offline_meeting_point["zip_code"]).strip()
                == online_meeting_point["zip_code"].strip()
            ):
                if (
                    offline_meeting_point["municipality"].strip().upper()
                    == online_meeting_point["name"].strip().upper()
                ) or (
                    offline_meeting_point["public_entry_address"].strip().upper()
                    == online_meeting_point["public_entry_address"].strip().upper()
                ):
                    is_online = True
                    print(
                        online_meeting_point["name"]
                        + " / "
                        + offline_meeting_point["municipality"]
                    )
                    break
                compare_distance = round(
                    distance.distance(
                        (
                            offline_meeting_point["latitude"],
                            offline_meeting_point["longitude"],
                        ),
                        (
                            online_meeting_point["latitude"],
                            online_meeting_point["longitude"],
                        ),
                    ).m,
                    2,
                )
                if compare_distance < 250:
                    is_online = True
                    print(
                        online_meeting_point["name"]
                        + " / "
                        + offline_meeting_point["municipality"]
                    )
                    break

        if not is_online:
            filtered_offline_meeting_points.append(offline_meeting_point)
    print("------- Done loading offline meeting points -------")
    return filtered_offline_meeting_points


def get_all_offline_meeting_points() -> List[OfflineMunicipality]:
    """
    Get global offline meeting points from the settings object
    """
    offline_meeting_point_stored: List[
        OfflineMunicipality
    ] = get_settings().offline_meeting_point_list.copy()
    return offline_meeting_point_stored


def set_all_offline_meeting_points(meeting_points: List[OfflineMunicipality]):
    """
    Set global offline meeting points list in the settings object
    """
    get_settings().offline_meeting_point_list = meeting_points
